import os
import pandas as pd
from Bio import SeqIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the directory paths
input_dir = r"C:\Input"
output_file = os.path.join(input_dir, "Leucine_and_Serine_analysis_combined.xlsx")

# Leucine and Serine codons
leucine_codons = ["TTA", "TTG", "CTT", "CTC", "CTA", "CTG"]
serine_codons = ["TCT", "TCC", "TCA", "TCG", "AGT", "AGC"]

# Leucine classifications
leucine_classifications = {
    "1-2-stop": ["TTA", "TTG"],
    "2-2-stop": ["CTA", "CTG"],
    "No-stop": ["CTT", "CTC"]
}

# Serine classifications
serine_classifications = {
    "1-2-stop": ["TCA", "TCG"],
    "2-2-stop": ["TCT", "TCC"],
    "No-stop": ["AGT", "AGC"]
}

def translate_sequence(seq):
    try:
        return str(seq.translate())
    except Exception as e:
        print(f"Error in translating sequence: {e}")
        return ""

def classify_and_count_codons(nucleotide_seq, codons, classifications):
    codon_counts = {key: 0 for key in classifications.keys()}
    total_count = 0
    for i in range(0, len(nucleotide_seq) - 2, 3):
        codon = nucleotide_seq[i:i + 3]
        if codon in codons:
            total_count += 1
            for classification, codon_list in classifications.items():
                if codon in codon_list:
                    codon_counts[classification] += 1
    return total_count, codon_counts

def process_all_fasta_files():
    all_results = []
    for filename in os.listdir(input_dir):
        if filename.endswith(".fasta") or filename.endswith(".fas"):
            fasta_path = os.path.join(input_dir, filename)
            print(f"Processing file: {filename}")
            try:
                records = SeqIO.parse(fasta_path, "fasta")
                for record in records:
                    nucleotide_seq = str(record.seq)
                    amino_acid_seq = translate_sequence(record.seq)

                    total_leucines, leucine_counts = classify_and_count_codons(nucleotide_seq, leucine_codons, leucine_classifications)
                    total_serines, serine_counts = classify_and_count_codons(nucleotide_seq, serine_codons, serine_classifications)
                    total_amino_acids = len(amino_acid_seq)

                    result = {
                        "Filename": filename,
                        "Record ID": record.id,
                        "Nucleotide Sequence": nucleotide_seq,
                        "Amino Acid Sequence": amino_acid_seq,
                        "Total Amino Acids": total_amino_acids,
                        "Total Leucines": total_leucines,
                        "Leucine 1-2-stop": leucine_counts["1-2-stop"],
                        "Leucine 2-2-stop": leucine_counts["2-2-stop"],
                        "Leucine No-stop": leucine_counts["No-stop"],
                        "Total Serines": total_serines,
                        "Serine 1-2-stop": serine_counts["1-2-stop"],
                        "Serine 2-2-stop": serine_counts["2-2-stop"],
                        "Serine No-stop": serine_counts["No-stop"],
                        "Issue": len(nucleotide_seq) % 3 != 0
                    }
                    all_results.append(result)
            except Exception as e:
                print(f"Error processing {filename}: {e}")

    df = pd.DataFrame(all_results)
    df.to_excel(output_file, sheet_name='Combined Analysis', index=False)

    workbook = load_workbook(output_file)
    sheet = workbook['Combined Analysis']
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in range(2, len(df) + 2):
        issue = sheet[f"O{row}"].value
        if issue:
            for col in range(1, len(df.columns) + 1):
                sheet.cell(row=row, column=col).fill = fill

    workbook.save(output_file)
    print(f"Combined analysis saved to {output_file}")

if __name__ == "__main__":
    process_all_fasta_files()
